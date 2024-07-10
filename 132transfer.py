import pandas as pd
import glob
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

def read_from_c4(filename):
    """Reads the Excel file starting from cell C4 and up to column AI."""
    try:
        # Read the Excel file starting from the 3rd row (0-based indexing, so 3 is the 4th row) and columns C to AI
        df = pd.read_excel(filename, skiprows=2, usecols="C:AI")
        return df
    except Exception as e:
        print(f"Error reading file {filename}: {e}")
        return pd.DataFrame()  # Return an empty DataFrame if reading fails

data_path = r"C:\Users\naird\OneDrive - Dun and Bradstreet\Documents\Communal Stuff\5-year forecast ISI\15 variables mirrors"
all_files = glob.glob(f"{data_path}/*.xls*")

# Output file path
output_path = 'C:/Users/naird/OneDrive - Dun and Bradstreet/Documents/132_CheckerC.xlsx'

# Create a new workbook for storing extracted data
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

# Set the starting cell for pasting data
start_row = 2  # Row 3 (to paste data from E3)
start_col = 5  # Column G (G is the 7th column)

# Variable to track the current row for pasting
current_row = start_row

for filename in all_files:
    # Read data starting from cell C4
    df = read_from_c4(filename)

    # Iterate through the DataFrame and paste data starting from E3
    for row in df.itertuples(index=False, name=None):
        for col_index, cell_value in enumerate(row, start=start_col):
            ws.cell(row=current_row, column=col_index, value=cell_value)
        current_row += 1
    # Add a gap after each file's data
    # current_row += 1
start_num = 7
# Add column numbering in row 1 starting from G1
for i in range(start_num, start_num + df.shape[1]-2):
    ws.cell(row=1, column=i, value=i - start_num + 1)

# Save the workbook with the data pasted
wb.save(output_path)
print("Data pasted starting from cell E3 successfully!")

# Load the source workbook and sheet
src_workbook = load_workbook(r"C:\Users\naird\OneDrive - Dun and Bradstreet\Documents\Country List.xlsx")
src_sheet = src_workbook.active

# Load the target workbook and sheet
tgt_workbook = load_workbook(output_path)
tgt_sheet = tgt_workbook.active

# Define columns to copy (e.g., B, C)
columns_to_copy = ["B", "C"]

for col_letter in columns_to_copy:
    for row in range(1, src_sheet.max_row + 1):
        cell_value = src_sheet[f"{col_letter}{row}"].value
        tgt_sheet[f"{col_letter}{row}"].value = cell_value

# Save the target workbook
tgt_workbook.save(output_path)
print("Columns copied successfully!")

# Load the existing workbook
wb = load_workbook(output_path)

# Access the workbook properties
wb_props = wb.properties

# Add custom properties
wb_props.title = "Sensitivity Label"
wb_props.subject = "Commercial in Confidence"

# Add last modified date to cell A1
ws = wb.active
last_modified_date = datetime.now().strftime("%Y/%m/%d")
ws['A1'] = f"{last_modified_date}"

# Save the workbook
wb.save(output_path)
print("Sensitivity label added successfully!")
print("Last modified date added successfully!")
print("Successful", output_path)
