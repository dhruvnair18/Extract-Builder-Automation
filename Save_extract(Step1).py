import openpyxl
# import os

# Define the workbook path
workbook_path = r"I:\Data\Unity Extract Builder\Extract Builder.xlsm"

# Load the workbook
wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)

# Retrieve values from specific cells
vPreset = "Forecast"  # ["Manual","Forecast","Monthly","Preset3","Preset4","Preset5","Preset6"]
vUpdate = "Update"    # ["Update","Don't Update"]
print(vPreset)
print(vUpdate)

# Get the DE Inventory sheet and Date
de_inventory_sheet = wb['DE Inventory']
de_date_sheet = wb["Dates"]



# Find the column index for vPreset in the second row (header row)
header = [cell.value for cell in de_inventory_sheet[2]]
vCol = header.index(vPreset) + 1  # Adding 1 because openpyxl is 1-based for columns
print(header)

# Get the value for DE Folder
vba_sheet = wb['VBA']
vDE_Folder = vba_sheet['C6'].value

# Create a new workbook object
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# Adding headers for the new workbook
output_ws.append(["DE Name", "DE File", "Coverage", "DE Update", "DE Frequency"])

# Find the next available column and row in the new workbook
max_col = output_ws.max_column + 1
max_row = output_ws.max_row + 1

# Save the new workbook (optional at this stage)
output_path = r"I:\Data management\Extract Builder Files\Extract.xlsx"
output_wb.save(output_path)

print(f"Next available column: {max_col}")
print(f"Next available row: {max_row}")


#Dates sheet copy of marcro file

"""
# Define the rows and start column (L is the 12th column in Excel)
specific_rows = [5, 10, 15]  # Example list of specific rows to process
start_column = 12  # Column L

# Iterate over the specific rows
for vRow in specific_rows:
    for col in range(start_column, de_date_sheet.max_column + 1):
        # Read the value from the Dates sheet
        value_to_copy = de_date_sheet.cell(row=vRow, column=col).value
        
        # Write the value to the Extract sheet
        output_ws.cell(row=vRow, column=col, value=value_to_copy)
        
 """      
        
# Iterate over rows in DE Inventory sheet

specific_rows = [1,2,3]  # Example list of specific rows to process
start_column = 12  # Column L



for i in range(3, 1001):
    vCoverage = de_inventory_sheet.cell(row=i, column=vCol).value
    
    if vCoverage not in [0, '', None]:
        vDE_File = de_inventory_sheet.cell(row=i, column=4).value
        vDE_Name = de_inventory_sheet.cell(row=i, column=3).value
        vDE_Update = de_inventory_sheet.cell(row=i, column=8).value
        vDE_Freq = de_inventory_sheet.cell(row=i, column=5).value
        vDE_File_path = de_inventory_sheet.cell(row=i,column=6).value
        for vRow2 in specific_rows:
            for col in range(start_column,de_date_sheet.max_column+1):
                if vDE_Freq == "A":
                    vRow2 = 1
                elif vDE_Freq == "Q":
                    vRow2 = 2
                elif vDE_Freq == "M":
                    vRow2 = 3
                else:
                    vRow2 = None
        vDE_Date = de_date_sheet.cell(row=vRow2,column=col)
        # Print the values for demonstration purposes
        print(f"Row {i}:")
        print(f"  Coverage: {vCoverage}")
        print(f"  DE File: {vDE_File}")
        print(f"  DE Name: {vDE_Name}")
        print(f"  DE Update: {vDE_Update}")
        print(f"  DE Frequency: {vDE_Freq}")
        print(f"  DE Path:{vDE_File_path}")
        # Determine vRow2 based on vDE_Freq
        

        # Append data to the next available row in the new workbook
        output_ws.append([vDE_Name, vDE_File, vCoverage, vDE_Update, vDE_Freq,vDE_File_path])

        # (Optional) Open the DE file
        """if vDE_File:
            file_path = os.path.join(vDE_Folder, vDE_File)
            try:
                print(f"Opening file: {file_path}")
                os.startfile(file_path)  # Opens the file
            except Exception as e:
                print(f"Failed to open file: {file_path}. Error: {e}")"""

# Save the final output workbook
output_wb.save(output_path)

#######

# Define the file path and sheet name
file_path = r"I:\Data management\Extract Builder Files\Extract.xlsx"
sheet_name = 'Sheet'  # Replace with your sheet name

# Load the workbook and select the sheet
wb = openpyxl.load_workbook(file_path)
sheet = wb[sheet_name]

# Get the last row number in the sheet
last_row = sheet.max_row

# Iterate through rows starting from row 2 to the last row
for row in range(2, last_row + 1):
    value = sheet.cell(row=row, column=6).value
    print(f'{value}')

