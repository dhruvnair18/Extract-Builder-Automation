import openpyxl
import os

# Source file path and sheet name
source_file_path = r"I:\Data management\Extract Builder Files\Extract.xlsx"
source_sheet_name = 'Sheet'  # Replace with your sheet name
source_column_index = 6  # Column index (A=1, B=2, C=3, ...)

# Directory containing files to match against
files_directory = r"I:\Data\All Data Elements"

# Load the source workbook and select the sheet
wb_source = openpyxl.load_workbook(source_file_path, read_only=True)
sheet_source = wb_source[source_sheet_name]

# Get the last row number in the source sheet
last_row_source = sheet_source.max_row

# List to store names from the source column
names_to_match = []

# Iterate through rows in the source sheet to get names
for row in range(1, last_row_source + 1):
    name = sheet_source.cell(row=row, column=source_column_index).value
    if name:
        names_to_match.append(name)

# Close the source workbook after reading
wb_source.close()

# Create a new workbook for storing extracted data
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Consolidated Data"

# Set the starting row for the new workbook
current_row = 1

# Add headers for the first two columns
new_sheet.cell(row=current_row, column=1, value='Filename')
new_sheet.cell(row=current_row, column=2, value='DE Name')
current_row += 1

# Maximum number of rows to extract from each file
max_rows_per_file = 230  # Adjust this number as per your requirement

# Iterate through each name and find matching files in the directory
for name in names_to_match:
    # Construct the file path based on the name
    file_path_to_open = os.path.join(files_directory, name)

    # Check if the file exists
    if os.path.exists(file_path_to_open):
        # Determine the file extension
        file_extension = os.path.splitext(file_path_to_open)[1].lower()

        try:
            # Check if the file is an xlsx or xlsm and open the appropriate sheet
            if file_extension == ".xlsx":
                wb_match = openpyxl.load_workbook(file_path_to_open, data_only=True)  # Use data_only=True to get the value of the formulas
                specific_sheet_name = 'Data'
            elif file_extension == ".xlsm":
                wb_match = openpyxl.load_workbook(file_path_to_open, data_only=True, keep_vba=True)  # Use data_only=True to get the value of the formulas
                specific_sheet_name = 'Data'
            else:
                print(f"Unsupported file type '{file_extension}' for file: {name}")
                continue  # Skip unsupported file types

            # Special case for files that need the label 'Commercial in Confidence'
            if name in ["DE188A", "DE189A", "DE1007A"]:
                specific_sheet_name = 'Commercial in Confidence'

            sheet_match = wb_match[specific_sheet_name]
            inputs_sheet = wb_match['Inputs']
            de_name = inputs_sheet['D4'].value  # Get DE Name from 'Inputs' sheet cell D4

            print(f"Opened file: {name}, extracting data from sheet: {specific_sheet_name}")

            # Initialize row counter for the current file
            rows_extracted = 0

            # Iterate through rows in the matched file to extract data
            for row in sheet_match.iter_rows():
                data_to_append = [cell.value for cell in row]  # Extract cell values (data_only=True ensures only values are read)
                # Insert Filename and DE Name in the first two columns
                new_sheet.cell(row=current_row, column=1, value=name)
                new_sheet.cell(row=current_row, column=2, value=de_name)
                for col_num, cell_value in enumerate(data_to_append, start=3):
                    new_sheet.cell(row=current_row, column=col_num, value=cell_value)
                current_row += 1
                rows_extracted += 1

                # Check if the maximum rows limit is reached
                if rows_extracted >= max_rows_per_file:
                    break

            # Close the matched workbook
            wb_match.close()

        except KeyError:
            print(f"Sheet '{specific_sheet_name}' or 'Inputs' not found in the file: {name}")
            continue  # Skip this file if the specific sheet is not found

    else:
        print(f"File '{name}' not found in the directory.")

# Delete columns from L (12th column) to BE (57th column) after extraction
columns_to_delete_start_L_BE = 12
columns_to_delete_end_L_BE = 57

for col in range(columns_to_delete_end_L_BE, columns_to_delete_start_L_BE - 1, -1):
    new_sheet.delete_cols(col)

# Find the last column with data
last_column = new_sheet.max_column

# Delete columns from BR (70th column) to the last column
columns_to_delete_start_BR = 27

for col in range(last_column, columns_to_delete_start_BR - 1, -1):
    new_sheet.delete_cols(col)

# Save the new workbook with the consolidated data
consolidated_file_path = r"I:\Data management\Extract Builder Files\Consolidated_Files\Consolidated_Data.xlsx"
new_workbook.save(consolidated_file_path)
print(f"Consolidated data saved to: {consolidated_file_path}")



import pandas as pd

# Load the Excel file and CSV file
file_path = r"I:\Data management\Extract Builder Files\Consolidated_Files\Consolidated_Data.xlsx"
csv_file_path = r"I:\Data management\Extract Builder Files\country_tiers.csv"

# Read the Excel file
df = pd.read_excel(file_path, sheet_name=None)

# Load the specific sheet into a DataFrame
sheet_name = 'Consolidated Data'
data = df[sheet_name]

# Load the data again with the correct headers starting from the second row
data = pd.read_excel(file_path, sheet_name=sheet_name, header=2)

# Rename the columns
data.columns.values[0] = 'Frequency'
data.columns.values[1] = 'DE Name'
# data.columns.values[7] = 'Date'
# Update the 'Frequency' column values to be the character before the '.'
data['Frequency'] = data['Frequency'].apply(lambda x: str(x).split('.')[0][-1] if pd.notnull(x) and '.' in str(x) else x)

# Delete the column with the heading "x"
data = data.drop(columns=['x'], errors='ignore')

# Select the columns to unpivot (L to X, which are columns 11 to 23 in 0-indexed pandas)
columns_to_unpivot = data.columns[12:25]

# Unpivot the specified columns
unpivoted_data = pd.melt(data, id_vars=data.columns[:10], value_vars=columns_to_unpivot, var_name='Year', value_name='Value')

# Read the CSV file containing ISO_2char information
iso_data = pd.read_csv(csv_file_path)

# Ensure that 'Namibia' has 'NA' as ISO_2char
iso_data.loc[iso_data['Countries'] == 'Namibia', 'ISO_2char'] = 'NA_'

# Merge the unpivoted data with the ISO data based on the country name
merged_data = unpivoted_data.merge(iso_data[['Countries', 'ISO_2char']], left_on='Country', right_on='Countries', how='left')

# Drop the 'Countries' column as it is redundant
merged_data = merged_data.drop(columns=['Countries'])

# Delete all rows where 'ISO_2char' is null
merged_data = merged_data.dropna(subset=['ISO_2char'])

# Rename the last column to 'ISO code' and move it to the 6th position
merged_data.rename(columns={'ISO_2char': 'ISO2'}, inplace=True)
merged_data.rename(columns={'Year': 'Date'}, inplace=True)
merged_data.rename(columns={'Data Element Identifier': 'DEID'}, inplace=True)


# Create a new column 'DE' by taking 'DEID' and removing the ISO2 from the end
merged_data['DE'] = merged_data.apply(lambda row: str(row['DEID'])[:-(len(str(row['Country Code'])))], axis=1)

# Move the 'DE' column to the first position
columns = ['DE'] + [col for col in merged_data.columns if col != 'DE']
merged_data = merged_data[columns]

# Delete columns F to L
columns_to_delete = ['Currency', 'Units', 'Information Source', 'Data Provider', 'Base Year', 'URL', 'Notes']  # Replace with actual column names
merged_data = merged_data.drop(columns=columns_to_delete, errors='ignore')

# Filter out rows with empty 'Value' column
merged_data = merged_data.dropna(subset=['Value'])

# Move 'ISO2' and 'Frequency' columns to the desired positions
cols = merged_data.columns.tolist()
# Move 'ISO2' to the 6th position (index 5 in 0-based indexing)
cols.insert(6, cols.pop(cols.index('ISO2')))
# Move 'Frequency' to the 7th position (index 6 in 0-based indexing)
cols.insert(6, cols.pop(cols.index('Frequency')))
merged_data = merged_data[cols]

# Save the modified data to a new Excel file
output_file_path = r"I:\Data management\Extract Builder Files\Consolidated_Files\Unpivoted_Consolidated_Data.xlsx"
merged_data.to_excel(output_file_path, index=False)

print(f'Modified data saved to {output_file_path}')


# Load the Excel file
file_path = r"I:\Data management\Extract Builder Files\Consolidated_Files\Unpivoted_Consolidated_Data.xlsx"
excel_data = pd.ExcelFile(file_path)

# Load the data from the first sheet
data = pd.read_excel(file_path, sheet_name='Sheet1')

# Sort the data first by 'DE' (first column) and then by 'Country' (third column) in ascending order
sorted_data = data.sort_values(by=['DE', 'Country'], ascending=[True, True])

# Display the first few rows of the sorted data
print(sorted_data.head())

# Save the sorted data to a new Excel file if needed
sorted_data.to_excel(r"I:\Data management\Extract Builder Files\Consolidated_Files\Sorted_Unpivoted_Consolidated_Data.xlsx", index=False)

output_file_path_txt = r"I:\Data management\Extract Builder Files\Text_Extracts\Sorted_Unpivoted_Consolidated_Data.txt"
merged_data.to_csv(output_file_path_txt, index=False, sep='\t')

